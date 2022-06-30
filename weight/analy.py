import numpy as np
import math
import pickle
import openpyxl as op

key_word = '文案编辑'
column = None


def length_fun(n: int):
    return n


def position_fun(p: float):
    return p


def result_fun(n: int, p: float):
    return n * p


def reset_words(words: list, polymerize: dict):
    polymerize_set = set([x for i in polymerize.values() for x in i])
    for i in range(len(words)):
        for j in range(len(words[i])):
            word = words[i][j]
            if word in polymerize_set:
                for _key, _value in polymerize.items():
                    if word in _value:
                        words[i][j] = _key
                        break


with open('../testing_data/result/{}.data'.format(key_word), 'rb') as f:
    abilities = pickle.load(f)
with open('./polymerize.data', 'rb') as f:
    polymerize = pickle.load(f)
reset_words(abilities, polymerize)

count_pos = []
for words in abilities:
    item = []
    for i, word in enumerate(words):
        word = word.replace('编程', '').replace('能力', '').lower()
        if len(words) == 1:
            item.append(0)
        else:
            item.append(i / (len(words) - 1))
    count_pos.append(item)

count_index = {}
for words, poses in zip(abilities, count_pos):
    for word, pos in zip(words, poses):
        word = word.replace('编程', '').replace('能力', '').lower()
        if word not in count_index:
            count_index[word] = []
        count_index[word].append(pos)

count_mean = {key: np.mean(value) for key, value in count_index.items()}
count_len = {key: len(value) for key, value in count_index.items()}
with open('./std_factor_细分.data', 'rb') as f:
    std_fact = pickle.load(f)
results = {key: result_fun(length_fun(count_len[key]), position_fun(1 - count_mean[key])) for key in count_len.keys()}
results = {key: value * std_fact[key] for key, value in results.items() if key in std_fact}
index = np.argsort(list(results.values()))[-1::-1]

workbook = op.load_workbook('../0627测试结果.xlsx')
worksheet = workbook['Sheet3']
if not column:
    column = 1
    while worksheet.cell(1, column).value:
        column = column + 2
worksheet.cell(1, column).value = key_word
worksheet.merge_cells(start_row=1, end_row=1, start_column=column, end_column=column+1)
for i in range(20):
    worksheet.cell(i+2, column).value = list(results.keys())[index[i]]
    worksheet.cell(i+2, column + 1).value = list(results.values())[index[i]]
workbook.save('../0627测试结果.xlsx')
